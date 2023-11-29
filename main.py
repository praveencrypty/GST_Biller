import os
import datetime
import customtkinter as ct
import openpyxl as xl
import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows

# lists to store product data
products = []
labels = []
quants = []
price = []

months = ('January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November',
          'December')

pathSale = None
pathPurchase = None

rowNumber = 5

date = datetime.date.today()
month = datetime.datetime.now().month
year = datetime.datetime.now().year

currentMonth = months[month-1];

print("Current Date: ",date)
print("Current month: ", currentMonth)



customerDetails = {
    "Date": "10/10/10",
    "Name": "Customer Name",
    "CustomerGST": "0000000000000",
    "Email": "samplemail@email.com",
    "Address": "Customer Address",
    "Phone": "9876543210",
}

fullData = None
selectedWorksheet = None
customerName = None
customerEmail = None
date = None
customerGST = None
customerAddress = None

submitButton = None

def button_callback():
    global fullData
    storeData()
    fullData = pd.DataFrame([customerDetails])

    print(fullData)
    OpenWorksheetToEdit()

def storeData():
    global customerDetails
    customerDetails['Name'] = customerName.get()
    customerDetails['Email'] = customerEmail.get()
    customerDetails['Phone'] = customerPhone.get()
    customerDetails['CustomerGST'] = customerGST.get()
    customerDetails['Address'] = customerAddress.get()
    customerDetails['Date'] = date.get()



def OpenWorksheetToEdit():
    global selectedWorksheet

    currentWorksheetName = str(currentMonth) + ".xlsx"
    toOpen = os.getcwd()
    toOpen = os.path.join(toOpen, "Bills/Sale/")
    toOpen = os.path.join(toOpen, currentWorksheetName)
    selectedWorksheet = xl.load_workbook(toOpen)

    #The Below will check the last row value and then add the data frame values to the next row in excel sheet
    ws = selectedWorksheet.active
    length = len(ws['A'])
    print(length)

    if length<=1:
        rows = dataframe_to_rows(fullData, index=False, header=True)
    else:
        rows = dataframe_to_rows(fullData, index=False, header=False)
        length += 1

    for r_idx, row in enumerate(rows, length):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    selectedWorksheet.save(toOpen)


def setZoomed():
    app.state('zoomed')


def createButton(R, tabname):
    global submitButton
    submitButton = ct.CTkButton(master=tabview.tab(tabname), text='Submit', command=button_callback)
    submitButton.grid(row=R, column=0, columnspan=5, padx=20, pady=20)


def deleteButton():
    global submitButton
    submitButton.destroy()


# Checkbox value once changed this function will be called
def checkbox_callback(value):
    global products
    global labels
    global quants
    global price

    length = len(products)
    for l in range(length):
        products[l].destroy()
        labels[l].destroy()
        quants[l].destroy()
        price[l].destroy()

    # Clear all stored data to restore new data
    products.clear()
    labels.clear()
    quants.clear()
    price.clear()

    total = int(value)
    for x in range(total):
        createProductDesc(x)



def appInit():
    global pathSale, pathPurchase

    currentDir = os.getcwd()
    path = os.path.join(currentDir, "Bills")

    if os.path.exists(path) == False:
        os.mkdir(path)

        pathSale = os.path.join(path, "Sale")
        pathPurchase = os.path.join(path, "Purchase")

        os.mkdir(pathSale)
        os.mkdir(pathPurchase)

        for m in months:
            createWorksheets(months.index(m), pathSale)
            createWorksheets(months.index(m), pathPurchase)




def createWorksheets(x, path):
    workBook = xl.Workbook()
    filename = str(months[x]) + ".xlsx"
    WorkbookPath = os.path.join(path, filename)
    workBook.save(WorkbookPath)


# Create The Product Description and Quantity GUI
def createProductDesc(rowVal):
    label = ct.CTkLabel(master=tabview.tab("Sale"), text=rowVal + 1, height=35)
    label.grid(row=rowNumber + rowVal, column=0, padx=2, pady=2, sticky="NSEW")
    tabview.grid_columnconfigure(0, weight=1)

    prod = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Product Description", state="normal", width=800, height=35)
    prod.grid(row=rowNumber + rowVal, column=1, padx=2, pady=2, sticky="NSEW")
    tabview.grid_columnconfigure(1, weight=1)

    quantity = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Quantity", state="normal", width=200, height=35)
    quantity.grid(row=rowNumber + rowVal, column=2, padx=2, pady=2, sticky="NSEW")
    tabview.grid_columnconfigure(2, weight=1)

    priceBar = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Price", state="normal", width=200, height=35)
    priceBar.grid(row=rowNumber + rowVal, column=3, padx=2, pady=2, sticky="NSEW")
    tabview.grid_columnconfigure(2, weight=1)

    labels.append(label)
    products.append(prod)
    quants.append(quantity)
    price.append(priceBar)


# ---------------------------------------------------------------------#
# MAIN Below


appInit()

app = ct.CTk()


app.title("GST Biller")
app.geometry("1920x1080")

mainframe = ct.CTkFrame(app)
mainframe.pack(padx=50, pady=50, expand = True)

app.after(2, setZoomed)

# Create and show Tabs View
tabview = ct.CTkTabview(master=mainframe)
tabview.grid(padx=50, pady=50,  sticky="NSEW")



# Add Tabs to the tabview
tabview.add("Sale")  # add tab at the end
tabview.add("Purchase")  # add tab at the end

tabview.set("Sale")  # set currently visible tab


# __________Create Customer Info Section

customerName = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Customer name", state="normal", height=50)
customerName.grid(row=0, column=0, padx=10, pady=10, sticky="NSEW", columnspan=2)



customerEmail = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Customer Email", state="normal", height=50)
customerEmail.grid(row=0, column=2, padx=10, pady=10, sticky="NSEW")

date = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="DD/MM/YY", state="normal", height=50)
date.grid(row=0, column=3, padx=10, pady=10, sticky="NSEW")

customerGST = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Customer GST", state="normal", height=50)
customerGST.grid(row=1, column=3, padx=10, pady=10,sticky="NSEW")

customerAddress = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Customer Address", state="normal", height=50)
customerAddress.grid(row=1, column=0, padx=10, pady=10, sticky="NSEW", columnspan = 2)

customerPhone = ct.CTkEntry(master=tabview.tab("Sale"), placeholder_text="Customer Phone", state="normal", height=50)
customerPhone.grid(row=1, column=2, padx=10, pady=10, sticky="NSEW")

# __________Create Customer Info Section End

numlabel = ct.CTkLabel(master=tabview.tab("Sale"), text="No. Of Products")
numlabel.grid(row=2, column=0, padx=10, pady=20, sticky="NSEW")

combobox = ct.CTkComboBox(master=tabview.tab("Sale"), values=["1", "2", "3", "4", "5"], command=checkbox_callback)
combobox.set("1")
combobox.grid(row=2, column=1, padx=10, pady=20, sticky="NSEW")

checkbox_callback(combobox.get())

createButton(10, "Sale")

button = ct.CTkButton(master=tabview.tab("Purchase"), text="my button", command=button_callback)
button.grid(padx=20, pady=20)




# Main Loop will start the program
app.mainloop()

